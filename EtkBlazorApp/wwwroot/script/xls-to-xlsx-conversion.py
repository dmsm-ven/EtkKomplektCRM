import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook

# content is a string containing the file. For example the result of an http.request(url).
# You can also use a filepath by calling "xlrd.open_workbook(filepath)".

xlsBook = xlrd.open_workbook(file_contents=content)
workbook = openpyxlWorkbook()

for i in xrange(0, xlsBook.nsheets):
    xlsSheet = xlsBook.sheet_by_index(i)
    sheet = workbook.active if i == 0 else workbook.create_sheet()
    sheet.title = xlsSheet.name

    for row in xrange(0, xlsSheet.nrows):
        for col in xrange(0, xlsSheet.ncols):
            sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

# The new xlsx file is in "workbook", without iterators (iter_rows).
# For iteration, use "for row in worksheet.rows:".
# For range iteration, use "for row in worksheet.range("{}:{}".format(startCell, endCell)):".